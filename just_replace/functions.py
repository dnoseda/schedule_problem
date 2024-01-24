def adhoc_distance(newlist):
    """
    check distance by sum of the differences with original pos in whatever rotation
    """
    distance = 0
    for i, dev in enumerate(newlist):
        if is_MLB_group(dev):
            continue

        original_pos = int(dev[0:2]) % half_point
        new_pos = i % half_point
        distance = distance + abs(original_pos - new_pos)
    return distance

def get_success_fitness():
    r = 0
    for i in range(max_len):
        r2pos = i % len(rota2)    
        r1pos = i % len(rota1)
        dev1 = rota1[r1pos]
        dev2 = rota2[r2pos]
        if is_adjacent(rota1,rota2,i,i):
            print("<<<< add 70 is adjacent {} {}".format(dev1, dev2))
            r +=70
            print("><<<< r: ",r)
        if is_same_boss(rota1[r1pos], rota2[r2pos]):
            r +=1
        if are_both_new(rota1[r1pos], rota2[r2pos]):
            r +=1

        if i <= max_len/2 and (dev1 in LAST_MONTH_L or dev2 in LAST_MONTH_L):
            r +=1
    print(">>> R",r)
    return r #+ adhoc_distance(rota1+rota2)

def get_boss(cel):
    return cel[2]

def is_MLB_group(dev):
    return dev[0] == "G"

def is_new(cel):    
    return cel[-1:] == "x" or is_MLB_group(cel)

def check_mlb_adjacent(rota1, rota2, i,j):
    """
    i = current position
    j = future position
    """

    # check that rota2 dev is going to be ok with future pos in J

    
    future_r2_next = rota2[(j+1) % len(rota2)]
    future_r2_prev = rota2[(j-1) % len(rota2)]

  
    

    next_to_group_pos = 0
    prev_to_group_pos = 0
    sec_mlb_group = 0

    # in next pos mlb group can only be adjacent to another mlb group
    if is_MLB_group(future_r2_next):
        
        next_to_group_pos = j + 2
        prev_to_group_pos = j - 1
        sec_mlb_group = j+1
        #print("Means pos " ,future_r2_next, " is next to mlb group in pos",sec_mlb_group)
        if is_MLB_group(future_r2_prev):
            return True
    if is_MLB_group(future_r2_prev):
        next_to_group_pos = j + 1
        prev_to_group_pos = j - 2
        sec_mlb_group = j-1
        #print("Means pos " ,future_r2_prev, " is next to mlb group in pos",sec_mlb_group)
        if is_MLB_group(future_r2_next):
            return True
    else:
        return True # no mlb group adjacent to next pos
    
    future_r2_next_to_group = rota2[next_to_group_pos % len(rota2)]
    future_r2_prev_to_group = rota2[prev_to_group_pos % len(rota2)]

    future_r1_next_to_group = rota1[next_to_group_pos % len(rota1)]
    future_r1_prev_to_group = rota1[prev_to_group_pos % len(rota1)]

    future_r1_pos_1 =  rota1[prev_to_group_pos+1 % len(rota1)]
    future_r1_pos_2 =  rota1[prev_to_group_pos+2 % len(rota1)]
    
    
    adjacent_bosses = [future_r2_next_to_group, future_r2_prev_to_group, future_r1_next_to_group, future_r1_prev_to_group, future_r1_pos_1, future_r1_pos_2]

    # FIXME hay que chequear si queremos que los grupos adjacentes sean siempre con mismo dev o diferente

    # liders de los 2 grupos
    #print("sec_mlb_group is ", sec_mlb_group)
    g1, g2 = rota2[i % len(rota2)], rota2[sec_mlb_group % len(rota2)]
    mlb_group_leads = mlb_group_lead[g1] + mlb_group_lead[g2]

    # check intersecciones
    for boss in mlb_group_leads:
        if boss in adjacent_bosses:
            return True
        
    return False


def is_adjacent(rota1, rota2, i, j):
    """
    i = current position
    j = future position
    """
    devR2 = rota2[i % len(rota2)]

    if is_MLB_group(devR2):
        ret= check_mlb_adjacent(rota1, rota2, i,j)
        # print all incomming parameters
        print("MLB Is adjacent {}, rota1 dev [{}]: {} rota2 dev[{}] {}, next {} rota1 {}, next {} rota2 {}".format(
                ret,
                i,rota1[i%len(rota1)],
                i,rota2[i%len(rota2)],
                j,rota1[j%len(rota1)],
                j,rota2[j%len(rota2)]
            ))
        return ret




    a1, a2 = get_boss(rota1[i % len(rota1)]), get_boss(rota1[(j+1)%len(rota1)])
    b1, b2 = get_boss(rota2[i % len(rota2)]), get_boss(rota2[(j+1)%len(rota2)])

     

    if a1 in [a2, b2]:
        return True            
    
    if b1 in [b2, a2]:
        return True

    return False

def is_same_boss(dev1, dev2):
    if is_MLB_group(dev1) or is_MLB_group(dev2):
        return False
    return get_boss(dev1) == get_boss(dev2)
def are_both_new(dev1,dev2):
    return is_new(dev1) and is_new(dev2)
def is_in_last_month(dev):
    return dev in LAST_MONTH_L


def get_offset():
    #global last_offset
    #last_offset = last_offset + 1
    return random.randint(0, max_len) 
    #return last_offset
    

def replace_better_pos(rota1, rota2, i):
    """
    i = current position in both rota1 and rota2
    """
    offset = get_offset()
    for j in range(max_len):
        """
        j is next position
        """
        r2pos = (j+i+offset) % len(rota2)
        r1pos = (j+i+offset) % len(rota1)
       
        if not is_same_boss(rota1[i%len(rota1)], rota2[r2pos]) and not is_adjacent(rota1,rota2,i,j) and not are_both_new(rota1[i%len(rota1)], rota2[r2pos]):
            #replace
            bf = get_success_fitness()
            
            rota1[i%len(rota1)], rota2[r2pos] = rota2[r2pos], rota1[i%len(rota1)]
            af = get_success_fitness()
            if bf >= af:
                #print("Switch : ",rota1[i%len(rota1)], "->", rota2[r2pos], "from", i, "to", r2pos)
                return rota1[i%len(rota1)], rota2[r2pos]
            else: #rollback                
                rota1[i%len(rota1)], rota2[r2pos] = rota2[r2pos], rota1[i%len(rota1)]

        if not is_adjacent(rota1,rota2,i,j) and not are_both_new(rota1[i%len(rota1)], rota2[r2pos]):
            
            bf = get_success_fitness()
            rota1[r1pos], rota1[i%len(rota1)] = rota1[i%len(rota1)], rota1[r1pos]            
            af = get_success_fitness()
            if bf >= af:
                #print("Switch : ",rota1[r1pos], "->", rota1[r1pos], "from", i, "to", r1pos)
                return rota1[i%len(rota1)], rota2[r2pos]
            else: #rollback                
                rota1[r1pos], rota1[i%len(rota1)] = rota1[i%len(rota1)], rota1[r1pos]
        
        
    
    return "",""
def write_solution_to_excel(rota1, rota2, nrota1, nrota2, people_dict):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    
    # Create a new workbook
    workbook = Workbook()
    sheet = workbook.active
    
    sheet['A1'] = "New rotation"    

    sheet['A2'] = "Original:"
    sheet.merge_cells('A2:D2')

    sheet["F2"] = "New:"
    sheet.merge_cells('F2:I2')

    sheet['A3'], sheet['C3'], sheet['F3'], sheet['H3'] = "Rota1", "Rota2", "Rota1", "Rota2"
    sheet.merge_cells('A3:B3')
    sheet.merge_cells('C3:D3')
    sheet.merge_cells('F3:G3')
    sheet.merge_cells('H3:I3')

    for i in ["A","C", "F", "H"]:
        sheet[i+"4"] = "Name"
    for i in ["B","D", "G", "I"]:
        sheet[i+"4"] = "Leader"

    sheet["E4"] = "Is Adjacent?"
    sheet["J4"] = "Is Adjacent?"

    titles=  ["A1","A2", "F2","A3", "C3", "F3","H3","A4","C4", "F4", "H4", "B4","D4", "G4", "I4", "E4", "J4"]       

    offset = 5


    """
    no cambiar cada 8 guardias
    2 grupos de mlb con solo 2 personas
    Habría GA1 y GA2
    Siempre consecutivos
    los grupos de MLB se agregan al final de la lista
    los devs de esos grupos se eliminan de la lista, no hay replace
    """

    for i in range(max(len(rota1), len(rota2))):
        pos = i + offset
        odevr1, odevr2, ndevr1, ndevr2 = rota1[(i % len(rota1))], rota2[(i % len(rota2))], nrota1[(i % len(nrota1))], nrota2[(i % len(nrota2))]

        sheet["A"+str(pos)] = people_dict[odevr1]["name"]
        if people_dict[odevr1]["has_experience"] == "FALSE":
            sheet["A"+str(pos)].fill = PatternFill(start_color="FFFF00",end_color="FFFF00", fill_type="solid") # yellow
        elif is_in_last_month(odevr1):
            sheet["A"+str(pos)].fill = PatternFill(start_color="CCFFFF",end_color="CCFFFF", fill_type="solid") # light blue

        
        sheet["B"+str(pos)] = get_lead(odevr1)

        sheet["C"+str(pos)] = people_dict[odevr2]["name"]
        if people_dict[odevr2]["has_experience"] == "FALSE":
            sheet["C"+str(pos)].fill = PatternFill(start_color="FFFF00",end_color="FFFF00", fill_type="solid") # yellow
        elif is_in_last_month(odevr2):
            sheet["C"+str(pos)].fill = PatternFill(start_color="CCFFFF",end_color="CCFFFF", fill_type="solid") # light blue

        sheet["D"+str(pos)] = get_lead(odevr2)
        
        sheet["E"+str(pos)] = "=OR(D{pos}=D{npos},D{pos}=B{npos},B{pos}=B{npos},B{pos}=D{npos})".format(pos=pos,npos=(pos+1))

        ### New:
        sheet["F"+str(pos)] = people_dict[ndevr1]["name"]
        if people_dict[ndevr1]["has_experience"] == "FALSE":
            sheet["F"+str(pos)].fill = PatternFill(start_color="FFFF00",end_color="FFFF00", fill_type="solid") # yellow
        elif is_in_last_month(ndevr1):
            sheet["F"+str(pos)].fill = PatternFill(start_color="CCFFFF",end_color="CCFFFF", fill_type="solid") # light blue

        sheet["G"+str(pos)] = get_lead(ndevr1)
    
        sheet["H"+str(pos)] = people_dict[ndevr2]["name"]
        if people_dict[ndevr2]["has_experience"] == "FALSE":
            sheet["H"+str(pos)].fill = PatternFill(start_color="FFFF00",end_color="FFFF00", fill_type="solid") # yellow
        elif is_in_last_month(ndevr2):
            sheet["H"+str(pos)].fill = PatternFill(start_color="CCFFFF",end_color="CCFFFF", fill_type="solid") # light blue

        sheet["I"+str(pos)] = get_lead(ndevr2)

        sheet["J"+str(pos)]="=OR(I{pos}=I{npos},I{pos}=G{npos},G{pos}=G{npos},G{pos}=I{npos})".format(pos=pos,npos=(pos+1))
    
    offset += max(len(rota1), len(rota2) )+ 2
    sheet["A"+str(offset)] = "Code list"
    titles.append("A"+str(offset))
    offset +=1    

    for index, key in enumerate(people_dict):  
        pos = index + offset
        sheet["A"+str(pos)] = key
        sheet["B"+str(pos)] = people_dict[key]['name']
        sheet["C"+str(pos)] = people_dict[key]['leader']
    
    offset += len(people_dict) + 2

    sheet["A"+str(offset)] = "Last month"
    titles.append("A"+str(offset))
    offset +=1
    for index, key in enumerate(LAST_MONTH_L):
        pos = index + offset
        sheet["A"+str(pos)] = key
        sheet["B"+str(pos)] = people_dict[key]['name']

    
    for cell in titles:
        sheet[cell].alignment = Alignment(horizontal='center')
        sheet[cell].font = Font(bold=True)
    
    workbook.save("report.xlsx")
    print("Excel file created successfully.")
        
