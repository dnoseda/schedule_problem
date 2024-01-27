import random

LAST_MONTH_L=[]

def write_solution_to_excel(rota1, rota2, nrota1, nrota2, people_dict):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    
    # Create a new workbook
    workbook = Workbook()
    sheet = workbook.active
    
    sheet['A1'] = "New rotation"    

    sheet['A2'] = "Original:"
    sheet.merge_cells('A2:D2')

    sheet["F2"] = "New:"
    sheet.merge_cells('F2:I2')

    sheet['A3'], sheet['C3'], sheet['F3'], sheet['H3'] = "Rota1", "Rota2", "Rota1", "Rota2"
    sheet.merge_cells('A3:B3')
    sheet.merge_cells('C3:D3')
    sheet.merge_cells('F3:G3')
    sheet.merge_cells('H3:I3')

    for i in ["A","C", "F", "H"]:
        sheet[i+"4"] = "Name"
    for i in ["B","D", "G", "I"]:
        sheet[i+"4"] = "Leader"

    sheet["E4"] = "Is Adjacent?"
    sheet["J4"] = "Is Adjacent?"

    titles=  ["A1","A2", "F2","A3", "C3", "F3","H3","A4","C4", "F4", "H4", "B4","D4", "G4", "I4", "E4", "J4"]       

    offset = 5


    """
    no cambiar cada 8 guardias
    2 grupos de mlb con solo 2 personas
    Habría GA1 y GA2
    Siempre consecutivos
    los grupos de MLB se agregan al final de la lista
    los devs de esos grupos se eliminan de la lista, no hay replace
    """

    get_lead = lambda x: people_dict[x]["leader"]

    for i in range(max(len(rota1), len(rota2))):
        pos = i + offset
        odevr1, odevr2, ndevr1, ndevr2 = rota1[(i % len(rota1))], rota2[(i % len(rota2))], nrota1[(i % len(nrota1))], nrota2[(i % len(nrota2))]

        sheet["A"+str(pos)] = people_dict[odevr1]["name"]
        name_cell = sheet["A"+str(pos)]
        if people_dict[odevr1]["has_experience"] == "FALSE":
            name_cell.fill = PatternFill(start_color="FFFF00",end_color="FFFF00", fill_type="solid") # yellow
        elif is_in_last_month(odevr1):
            name_cell.fill = PatternFill(start_color="CCFFFF",end_color="CCFFFF", fill_type="solid") # light blue

        
        sheet["B"+str(pos)] = get_lead(odevr1) # FIXME implement get_lead, get leader name by code

        sheet["C"+str(pos)] = people_dict[odevr2]["name"]
        r2_name_cell = sheet["C"+str(pos)]
        if people_dict[odevr2]["has_experience"] == "FALSE":
            r2_name_cell.fill = PatternFill(start_color="FFFF00",end_color="FFFF00", fill_type="solid") # yellow
        elif is_in_last_month(odevr2):
            r2_name_cell.fill = PatternFill(start_color="CCFFFF",end_color="CCFFFF", fill_type="solid") # light blue

        sheet["D"+str(pos)] = get_lead(odevr2)
        
        sheet["E"+str(pos)] = "=OR(D{pos}=D{npos},D{pos}=B{npos},B{pos}=B{npos},B{pos}=D{npos})".format(pos=pos,npos=(pos+1))

        ### New:
        sheet["F"+str(pos)] = people_dict[ndevr1]["name"]
        if people_dict[ndevr1]["has_experience"] == "FALSE":
            sheet["F"+str(pos)].fill = PatternFill(start_color="FFFF00",end_color="FFFF00", fill_type="solid") # yellow
        elif is_in_last_month(ndevr1):
            sheet["F"+str(pos)].fill = PatternFill(start_color="CCFFFF",end_color="CCFFFF", fill_type="solid") # light blue

        sheet["G"+str(pos)] = get_lead(ndevr1)
    
        sheet["H"+str(pos)] = people_dict[ndevr2]["name"]
        if people_dict[ndevr2]["has_experience"] == "FALSE":
            sheet["H"+str(pos)].fill = PatternFill(start_color="FFFF00",end_color="FFFF00", fill_type="solid") # yellow
        elif is_in_last_month(ndevr2):
            sheet["H"+str(pos)].fill = PatternFill(start_color="CCFFFF",end_color="CCFFFF", fill_type="solid") # light blue

        sheet["I"+str(pos)] = get_lead(ndevr2)

        sheet["J"+str(pos)]="=OR(I{pos}=I{npos},I{pos}=G{npos},G{pos}=G{npos},G{pos}=I{npos})".format(pos=pos,npos=(pos+1))
    
    offset += max(len(rota1), len(rota2) )+ 2
    sheet["A"+str(offset)] = "Code list"
    titles.append("A"+str(offset))
    offset +=1    

    for index, key in enumerate(people_dict):  
        pos = index + offset
        sheet["A"+str(pos)] = key
        sheet["B"+str(pos)] = people_dict[key]['name']
        sheet["C"+str(pos)] = people_dict[key]['leader']
    
    offset += len(people_dict) + 2

    sheet["A"+str(offset)] = "Last month"
    titles.append("A"+str(offset))
    offset +=1
    for index, key in enumerate(LAST_MONTH_L):
        pos = index + offset
        sheet["A"+str(pos)] = key
        sheet["B"+str(pos)] = people_dict[key]['name']

    
    for cell in titles:
        sheet[cell].alignment = Alignment(horizontal='center')
        sheet[cell].font = Font(bold=True)
    
    workbook.save("report.xlsx")
    print("Excel file created successfully.")
        


rota1= ["00A1_","01B2_","02B3_","03A4_","04C5_","05C6_","06C7_","07D8_","08D9_","09D10_","10A11_","11A12_","12A13_","13E14_","14B15_","15F16_","16G17_","17G18_","18G19_","19H20_","20H21_","21G22_","22H23_","23H24_","24E25_","49I50x"]
rota2= ["25G26_","26B27_","27I28_","28C29_","29B30x","30D31x","31A32x","32A33x","33H34x","34J35_","35J36_","36E37_","37H38x","38F39_","39G40_","40F41x","41G42_","42F43_","43F44_","44D45_","45D46_","46D47_","47A48x","48H49x","50I51x"]
orota1, orota2 = [],[]
max_len = max(len(rota1), len(rota2))
l1,l2 = "", ""
last_offset = 1

def get_lead(dev):
    return dev[2:4]

def adhoc_distance(newlist):
    """
    check distance by sum of the differences with original pos in whatever rotation
    """
    distance = 0
    for i, dev in enumerate(newlist):
        if is_MLB_group(dev):
            continue

        original_pos = int(dev[0:2]) % half_point
        new_pos = i % half_point
        distance = distance + abs(original_pos - new_pos)
    return distance

def not_mlb_group_join(rota2,pos):
    dev = rota2[pos % len(rota2)]

    if not is_MLB_group(dev):
        return False    
    
    prev_dev = rota2[(pos-1) % len(rota2)]
    next_dev = rota2[(pos+1) % len(rota2)]

    if prev_dev == dev: #ok 
        if is_MLB_group(next_dev):
            return True # should not have next group
        
    
    if next_dev == dev:
        if is_MLB_group(prev_dev):
            return True ## should not have prev group
    
    return False

def same_boss_mlb_surround(rota1, rota2, i):

    # check whether rota2 is mlb group
    dev2 = rota2[i % len(rota2)]
    if not is_MLB_group(dev2):
        return False
    
    # identify where surrounding
    prev_dev2 = rota2[(i-1) % len(rota2)]
    next_dev2 = rota2[(i+1) % len(rota2)]

    surrounding_leads = []
    if next_dev2 == dev2: # same group
        surrounding_leads.append(get_boss(rota1[(i-1) % len(rota1)]))
        surrounding_leads.append(get_boss(rota1[(i+2) % len(rota1)]))

        surrounding_leads.append(get_boss(rota2[(i-1) % len(rota2)]))
        surrounding_leads.append(get_boss(rota2[(i+2) % len(rota2)]))
    elif prev_dev2 == dev2:
        surrounding_leads.append(get_boss(rota1[(i-2) % len(rota1)]))
        surrounding_leads.append(get_boss(rota1[(i+1) % len(rota1)]))
        
        surrounding_leads.append(get_boss(rota2[(i-2) % len(rota2)]))
        surrounding_leads.append(get_boss(rota2[(i+1) % len(rota2)]))


    # check for intersections of surrounding bosses rota2 and rota1
    for lead in surrounding_leads:
        if lead in mlb_group_lead[dev2]:
            return True
        
    return False


def get_success_fitness():
    r = 0
    
    for i in range(max_len):
        r2pos = i % len(rota2)    
        r1pos = i % len(rota1)
        dev1 = rota1[r1pos]
        dev2 = rota2[r2pos]

        # TODO: check that mlbgroups are adjacent with same ID
        
        eval_dict = {
            "adjacent": {"f":is_adjacent(rota1,rota2,i,i),"value":1},
            "same_boss": {"f":is_same_boss(dev1, dev2),"value":1},
            "both_new": {"f":are_both_new(dev1, dev2),"value":10},
            "last_month": {"f":i <= max_len/2 and (dev1 in LAST_MONTH_L or dev2 in LAST_MONTH_L),"value":1},
            "all_r2_mlb": {"f":is_MLB_group(dev1),"value":10}, # check all mlb groups on rota2 and none on rota1,
            "mlb_group_join": {"f":not_mlb_group_join(rota2,i),"value":10},
            "adj_mlb": {"f":same_boss_mlb_surround(rota1, rota2, i),"value":1},
        }

        for k,v in eval_dict.items():
            #print("¢¢¢¢¢¢¢{}: {}".format(k,v))
            if v["f"]:
                r += v["value"]
    
    return r #+ adhoc_distance(rota1+rota2)

def get_boss(cel):
    return cel[2]

def is_MLB_group(dev):
    return dev[0] == "G"

def is_new(cel):    
    return cel[-1:] == "x" or is_MLB_group(cel)


def is_adjacent(rota1, rota2, i, j):
    """
    i = current position
    j = future position
    """
    devR2 = rota2[i % len(rota2)]

    if is_MLB_group(devR2):
        return False


    a1, a2 = get_boss(rota1[i % len(rota1)]), get_boss(rota1[(j+1)%len(rota1)])
    b1, b2 = get_boss(rota2[i % len(rota2)]), get_boss(rota2[(j+1)%len(rota2)])

     

    if a1 in [a2, b2]:
        return True            
    
    if b1 in [b2, a2]:
        return True

    return False

def same_boss_MLB(dev1,dev2):
    if dev1 == dev2:
        return False
    
    leads1 = []
    if is_MLB_group(dev1):
        leads1 = mlb_group_lead[dev1]
    else:
        leads1 = [get_boss(dev1)]
    
    if is_MLB_group(dev2):
        leads2 = mlb_group_lead[dev2]
    else:
        leads2 = [get_boss(dev2)]

    # check intersecciones
    for boss in leads1:
        if boss in leads2:
            return True
    return False
    

def is_same_boss(dev1, dev2):
    if is_MLB_group(dev1) or is_MLB_group(dev2):
        return same_boss_MLB(dev1, dev2)
    return get_boss(dev1) == get_boss(dev2)
def are_both_new(dev1,dev2):
    return is_new(dev1) and is_new(dev2)
def is_in_last_month(dev):
    return dev in LAST_MONTH_L


def get_offset():
    #global last_offset
    #last_offset = last_offset + 1
    return random.randint(0, max_len) 
    #return last_offset
    

def replace_better_pos(rota1, rota2, i):
    """
    i = current position in both rota1 and rota2
    """
    offset = get_offset()
    for j in range(max_len):
        """
        j is next position
        """
        r2pos = (j+i+offset) % len(rota2)
        r1pos = (j+i+offset) % len(rota1)
       
        if not is_same_boss(rota1[i%len(rota1)], rota2[r2pos]) and not is_adjacent(rota1,rota2,i,j) and not are_both_new(rota1[i%len(rota1)], rota2[r2pos]):
            #replace
            bf = get_success_fitness()
            
            rota1[i%len(rota1)], rota2[r2pos] = rota2[r2pos], rota1[i%len(rota1)]
            af = get_success_fitness()
            if bf >= af:
                #print("Switch : ",rota1[i%len(rota1)], "->", rota2[r2pos], "from", i, "to", r2pos)
                return rota1[i%len(rota1)], rota2[r2pos]
            else: #rollback                
                rota1[i%len(rota1)], rota2[r2pos] = rota2[r2pos], rota1[i%len(rota1)]

        if not is_adjacent(rota1,rota2,i,j) and not are_both_new(rota1[i%len(rota1)], rota2[r2pos]):
            
            bf = get_success_fitness()
            rota1[r1pos], rota1[i%len(rota1)] = rota1[i%len(rota1)], rota1[r1pos]            
            af = get_success_fitness()
            if bf >= af:
                #print("Switch : ",rota1[r1pos], "->", rota1[r1pos], "from", i, "to", r1pos)
                return rota1[i%len(rota1)], rota2[r2pos]
            else: #rollback                
                rota1[r1pos], rota1[i%len(rota1)] = rota1[i%len(rota1)], rota1[r1pos]
        
        
    
    return "",""

import csv

people_dict ={}


filename_input = "people.csv"  # Replace with the name of your CSV file
leader_codes = {}  # Replace with the leader names and their codes

# Load the CSV file into a list of dictionaries
with open(filename_input, "r") as file:
    reader = csv.DictReader(file)
    people_list = list(reader)

# Convert the list of people into a dictionary with codes as keys

# Load the CSV file into a list of dictionaries
with open("last_month.csv", "r") as file:
    reader = csv.DictReader(file)
    last_month = list(reader)

devs = []
dev_by_name = {}
mlb_devs = []
for i, person in enumerate(people_list):
    if not (leader_codes.get(person["leader"]) != None):
        leader_codes[person["leader"]] = chr(len(leader_codes) + 65) # increment letter from quantity of leaders like autoinc
    
    # if it is  mlb the lead should be the same, another alternative is to do leader as a list of leaders
    has_experience = person["has_experience"] == "TRUE" or person["name"] in last_month
    code =  "{:02d}{}{}{}".format(
            i,
            leader_codes[person["leader"]],
            str(i+1),
            ("_" if has_experience else "x")
    )
    people_dict[code] = person
    dev_by_name[person["name"]] = code
    """
        if it is from mlb:
            add once as a single block
            set leader as all leaders from mlb, this should be setted aside
        else should keep the same logic
        
    """
    if person["mlb"] == "TRUE":
        mlb_devs.append(code)
    else:
        devs.append(code)

print(">>> Before turning mlb_devs into single blocks")
print(people_dict)

MAX_MLB_DEVS = 2



"""
if len(mlb_devs) % 2 != 0:
    #exit
    print("Error: mlb_devs_total is not even")
    exit(1)
"""


print(">>> MLB debs")
print(mlb_devs)

mlb_devs_groups ={} # dev name -> group code
mlb_group_lead ={} # group code -> leader code

group_id = 1
for i in range(len(mlb_devs)):
    
    group_code ="G_{:02d}".format(group_id)
    mlb_devs_groups[mlb_devs[i]] = group_code

    if mlb_group_lead.get(group_code) == None:
        mlb_group_lead[group_code] = [get_boss(mlb_devs[i])]
    else:
        if get_boss(mlb_devs[i]) not in mlb_group_lead[group_code]:
            mlb_group_lead[group_code].append(get_boss(mlb_devs[i]))

    if (i+1) % MAX_MLB_DEVS == 0:
        group_id += 1
    




print(">>> MLB Blocks")
print(mlb_devs_groups)
print(">>> MLB Group Leaders")
print(mlb_group_lead )



# add all keys from mlb_group_lead to devs at the end
devs.extend(list(mlb_group_lead.keys()))


# delete repeated from devs
devs = list(dict.fromkeys(devs))

# add again groups keys to set groups by pairs
devs.extend(list(mlb_group_lead.keys()))

print(">>> After turning mlb_devs into single blocks")
print(devs)

#exit(0)

original_devs_arrange = "-".join(devs)
half_point = int(len(devs)/2)

for lm in last_month:
    LAST_MONTH_L.append(dev_by_name[lm["name"]])

print(LAST_MONTH_L)
rota1, rota2 = devs[:half_point],devs[half_point:]
orota1, orota2 = rota1+[], rota2+[]
max_len = max(len(rota1), len(rota2))

# rota2 is the only one that can have mlb devs

max_repeat = 1000
for j in range(max_repeat):
    print("iter",j, "rate", get_success_fitness(),"adhoc_distance" ,adhoc_distance(rota1+rota2))    
    for i in range(max_len):
        r2pos = i % len(rota2)
        r1pos = i % len(rota1)

        """
            Check if it is mlb dev,
            for mlb dev, the same_boss has to check with 2 sets of leaders
            for mlb dev, it's treated as a single block that can be adjacent to other mlb dev. Consider this in same_boss
        """
        if is_same_boss(rota1[r1pos], rota2[r2pos]):
            l1,l2= replace_better_pos(rota1, rota2, i)
            continue
        if is_adjacent(rota1, rota2, i, i):
            l1,l2= replace_better_pos(rota1, rota2, i)
            continue
        if are_both_new(rota1[r1pos], rota2[r2pos]):
            l1,l2= replace_better_pos(rota1, rota2, i)
            continue
        if i <= max_len/2 and is_in_last_month(rota1[r1pos]):
            l1,l2 = replace_better_pos(rota1, rota2, i)
            continue
        if i <= max_len/2 and is_in_last_month(rota2[r2pos]):
            l1,l2 = replace_better_pos(rota1, rota2, i)
            continue
    if get_success_fitness() == 0:
        break

#write_solution_to_excel(orota1, orota2,rota1, rota2, people_dict)
    
def get_name(dev):
    if is_MLB_group(dev):
        return dev
    return people_dict[dev]["name"]

def get_name_detailed(dev):
    if is_MLB_group(dev):
        names = dev + ": "
        #iterate mlb_devs_groups and get the name of the dev
        for k, v in mlb_devs_groups.items():        
            if mlb_devs_groups[k] == dev:
                names += people_dict[k]["name"] + ", "
    # delete trailing ", "
        return names[:-2]
    else:
        return people_dict[dev]["name"]

        

def get_leader_by_code(code):
    # iterate leader_codes and find the code by value
    for k,v in leader_codes.items():
        if v == code:
            return k

def get_lead(dev):
    ret = ""
    if is_MLB_group(dev):
        for l in mlb_group_lead[dev]:
            ret += get_leader_by_code(l) + ", " # doesnt work, have to get list of leads of mlb_group_lead and then from
        # if ret containts ", " delete it
        if ", " in ret:
            ret = ret[:-2]
        return ret
    return people_dict[dev]["leader"]


print("\n>>> Fine tuning...\n")
## iterate and try replace rota2 cells with every position of rota2, then check if it's better than the current one with the get_success_fitness function
for i in range(len(rota2)):
    for j in range(len(rota2)):
        print("{:02d} - {:02d} = {}".format(i,j,get_success_fitness()))
        for r1i in range(len(rota1)):
            for r1j in range(len(rota1)):              
                if i == j:
                    continue
                before_perf = get_success_fitness()
                rota2[i], rota2[j] = rota2[j], rota2[i]
                rota1[r1i], rota1[r1j] = rota1[r1j], rota1[r1i]
                after_perf = get_success_fitness()
                if not(after_perf < before_perf):
                    rota2[i], rota2[j] = rota2[j], rota2[i]
                    rota1[r1i], rota1[r1j] = rota1[r1j], rota1[r1i]                    

print("\n>>> Leader codes\nName\tCode")
for k,v in leader_codes.items():
    print("{}\t{}".format(k,v))

print("\n>>> MLB Leaders group codes\nGroup\tLeaders")
for k,v in mlb_group_lead.items():    
    lead_group = ""
    
    for l in v:
        lead_group += ", ".join(map(get_leader_by_code, v))
    print("{}\t{}".format(k, lead_group))    


print("\n>>> Devs in mlb groups\nCode\tDev\tGroup")
for k,v in mlb_devs_groups.items():
    print("{}\t{}\t{}".format(k,people_dict[k]['name'],v))

print("\n>>> All Devs groups\nCode\tDev\tGroup\tleader")
for k,v in people_dict.items():
    print("{}\t{}\t{}\t{}".format(k,v['name'],v['mlb'],v['leader']))




print("\n>>> Devs in mlb groups\nCode R1\tCode R2\tadjacent\tsame boss\tboth new\tconflict?\tR1 Dev\tR1 Lead\tR2 Dev\tR2 Lead")
for i in range(max_len):
    r2pos = i % len(rota2)    
    r1pos = i % len(rota1)
    dev1 = rota1[r1pos]
    dev2 = rota2[r2pos]
    print("{} ({}) [{}]\t{} ({}) [{}]\t{}\t{}\t{}==\t{}\t\t{}{}\t{}\t{}{}\t{}".format(
        dev1, get_boss(dev1), dev1[-1],
        dev2, get_boss(dev2), dev2[-1],
        is_adjacent(rota1,rota2,i,i),
        is_same_boss(rota1[r1pos], rota2[r2pos]),
        are_both_new(rota1[r1pos], rota2[r2pos]),
        (
            is_adjacent(rota1,rota2,i,i) or
            is_same_boss(rota1[r1pos], rota2[r2pos]) or
            are_both_new(rota1[r1pos], rota2[r2pos])
        ),
        get_name_detailed(dev1),"(.)" if is_in_last_month(dev1) else "",get_lead(dev1),
        get_name_detailed(dev2),"(.)" if is_in_last_month(dev2) else "",get_lead(dev2),
        
        ))




