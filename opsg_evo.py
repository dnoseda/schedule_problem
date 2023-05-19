import random
import time
import sys
from multiprocessing import Pool
import csv
#from jinja2 import Template

#from openpyxl.formula import Formula

start = time.time()
    
def printT(msg):
    print((time.time()-start), "-->",msg)

POPULATION_SIZE = 1300
MUTATION_RATE = 0.5
MAX_GENERATIONS = 10000
RANDOM_SELECT = 0.5
MAX_CHILDS = int(POPULATION_SIZE * (1-RANDOM_SELECT))

LAST_MONTH_L = []


#devs = ['B1', 'D1x', 'B2', 'D2', 'C1', 'A2x', 'A1', 'C2']
devs = ['A01_','B02_','C03_','A04_','C05_','C06_','D07_','B08_','E09_','D10x','E11_','B12_','C13_','A14_','A15_','C16_','B17_','A18_','A19_','C20_','B21_','F22_','E23x','G24_','E25x','G26_','D27x','E28x','H29_','F30_','H31_','H32_','C33_','I34_','I35_','I36_','F37_','F38_','F39_','F40_','I41_','A42_','E43_','D44_','H45_','I46_']
original_devs_arrange = "-".join(devs)
half_point = int(len(devs)/2)

people_dict = {}


def printI(individual):
    print("Rota 1 -> ",end="")
    print(individual[:half_point])

    print("Rota 2 -> ",end="")
    print(individual[half_point:])

    print("CSV:")
    print("\t".join(individual[:half_point]))
    print("\t".join(individual[half_point:]))


filename = None

def save_solution(i, best_solution):
    csv_file = open(filename, 'a')
    writer = csv.writer(csv_file)
    row =  [
            i,
            fitness(best_solution), 
            adhoc_distance(best_solution)
        ]+best_solution
    
    writer.writerow(row)   
    csv_file.close()

def get_boss(cel):
    return cel[2]

def is_new(cel):
    return cel[-1:] == "x"

def is_adjacent(rota1, rota2, i):
    a1, a2 = get_boss(rota1[i]), get_boss(rota1[(i+1)%len(rota1)])
    b1, b2 = get_boss(rota2[i]), get_boss(rota2[(i+1)%len(rota2)])

    if a1 in [a2, b2]:
        return True            
    
    if b1 in [b2, a2]:
        return True

    #print("A1: {} A2 {} B1 {} B2 {} results in {}", a1, a2, b1, b2, res)
    
    return False


def adhoc_distance(newlist):
    """
    check distance by sum of the differences with original pos in whatever rotation
    """
    distance = 0
    for i, dev in enumerate(newlist):
        original_pos = int(dev[0:2]) % half_point
        new_pos = i % half_point
        distance = distance + abs(original_pos - new_pos)
    return distance
    

def fitness(individual, is_original=False):
    if not is_original and "-".join(individual)== original_devs_arrange:
        return 0.0000001
    
    conflicts = 0
    rota1, rota2 = individual[:half_point],individual[half_point:]
    fitness_contribution ={
        "same_boss": 0,
        "adjacent_boss": 0,
        "is_recent": 0
    }
    for i in range(half_point):

        # consider if it first time on schedule
        if is_new(rota1[i]) and is_new(rota2[i]):
            #print("NW")
            return 0.0000001

        # same dev
        if rota1[i] == rota2[i]:
            print("SD")
            return 0.0000001 #

        # same boss
        if get_boss(rota1[i]) == get_boss(rota2[i]): 
            fitness_contribution["same_boss"] += 1
        
        # adjacent boss
        if is_adjacent(rota1, rota2, i):
            fitness_contribution["adjacent_boss"] += 1        
         
        # discard solutions where first devs are from the last month
        """
        if i < (len(LAST_MONTH_L)/2):
            if is_in_last_month(rota1[i]) or is_in_last_month(rota2[i]):
                #print("LM")
                fitness_contribution["is_recent"] += 1
                """
    
    """
    fitness_contribution["distance"] =adhoc_distance(individual)
    """
    
    weights = {
        "same_boss": 100,
        "adjacent_boss": 1000,
        "is_recent": 1,
        "distance" : 1
    }

    for k, v in fitness_contribution.items():
        conflicts += weights[k] * ((v/500) if k == "distance" else v)
    
    #print("---->",fitness_contribution, " = ", conflicts)

    return (1 / (conflicts + 1))

def is_in_last_month(dev):
    return dev in LAST_MONTH_L
   
    

class EightQueensGA:
    def __init__(self):
        self.population = []
        self.fitness_scores = []
        self.best_solution = None
        self.csv_writer = None
        self.first_best= 0
        self.best_fitness = 0

        for i in range(POPULATION_SIZE):
            individual = random.sample(devs, len(devs))
            self.population.append(individual)
    def set_csv_writer(self, csv_writer):
        self.csv_writer = csv_writer

    def select_parents(self):
        parent1 = random.choices(self.population, weights=self.fitness_scores)[0]
        parent2 = random.choices(self.population, weights=self.fitness_scores)[0]
        return parent1, parent2

    def crossover(self, p1, p2):
        i = random.randint(0, len(p1) - 1)        
        #i = half_point
        
        child1 = p1[:i]
        child2 = p2[:i]        
        
        child1 = child1 + list(set(p2) - set(child1))
        child2 = child2 + list(set(p1) - set(child2))

        return child1, child2

    def mutate(self, individual):
        if random.random() < MUTATION_RATE:
            pos1, pos2 = random.sample(range(8), 2)
            individual[pos1], individual[pos2] = individual[pos2], individual[pos1]

    def evolve(self):
        for i in range(MAX_GENERATIONS):

            #self.fitness_scores = [fitness(individual) for individual in self.population]

            with Pool() as pool:
                self.fitness_scores = pool.map(fitness, self.population)

            sorted_devs = [x for _, x in sorted(zip(self.fitness_scores, self.population))]

            
            best_for_now = sorted_devs[-1]
            best_fit_for_now = fitness(best_for_now)
            
            if self.best_solution is None or self.best_fitness < best_fit_for_now:
                self.best_solution = best_for_now
                self.best_fitness = best_fit_for_now
                self.first_best = i
                save_solution(i, self.best_solution)                
                print("*")
                

            printT("start gen {} -> {} found on pop {} distance {}".format(
                i,
                self.best_fitness,
                self.first_best,
                adhoc_distance(self.best_solution)
                ))

            new_population = []

            

            for j in range(int(MAX_CHILDS / 2)):
                parent1, parent2 = self.select_parents()
                child1, child2 = self.crossover(parent1, parent2)
                self.mutate(child1)
                self.mutate(child2)
                new_population.append(child1)
                new_population.append(child2)

            while len(new_population) < POPULATION_SIZE:                
                individual = random.sample(devs, len(devs))
                new_population.append(individual)

            self.population = new_population

    def print_solution(self):
        print("Original:", fitness(devs, is_original=True))
        printI(devs)
        if self.best_solution is not None:
            print("Solution found with score.", fitness(self.best_solution), self.best_fitness)
            print("AdhocDist: ", adhoc_distance(self.best_solution))
            rota1, rota2 = devs[:half_point],devs[half_point:]
            nrota1, nrota2 = self.best_solution[:half_point],self.best_solution[half_point:]

            write_solution_to_excel(rota1, rota2, nrota1, nrota2, people_dict)
        else:
            print("No solution found.")

def write_solution_to_excel(rota1, rota2, nrota1, nrota2, people_dict):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    
    # Create a new workbook
    workbook = Workbook()
    sheet = workbook.active
    
    sheet['A1'] = "New rotation"    

    sheet['A2'] = "Original:"
    sheet.merge_cells('A2:D2')

    sheet["F2"] = "New:"
    sheet.merge_cells('F2:I2')

    sheet['A3'], sheet['C3'], sheet['F3'], sheet['H3'] = "Rota1", "Rota2", "Rota1", "Rota2"
    sheet.merge_cells('A3:B3')
    sheet.merge_cells('C3:D3')
    sheet.merge_cells('F3:G3')
    sheet.merge_cells('H3:I3')

    for i in ["A","C", "F", "H"]:
        sheet[i+"4"] = "Name"
    for i in ["B","D", "G", "I"]:
        sheet[i+"4"] = "Leader"

    sheet["E4"] = "Is Adjacent?"
    sheet["J4"] = "Is Adjacent?"

    titles=  ["A1","A2", "F2","A3", "C3", "F3","H3","A4","C4", "F4", "H4", "B4","D4", "G4", "I4", "E4", "J4"]       

    offset = 5

    for i in range(max(len(rota1), len(rota2))):
        pos = i + offset
        odevr1, odevr2, ndevr1, ndevr2 = rota1[(i % len(rota1))], rota2[(i % len(rota2))], nrota1[(i % len(nrota1))], nrota2[(i % len(nrota2))]

        sheet["A"+str(pos)] = people_dict[odevr1]["name"]
        if people_dict[odevr1]["has_experience"] == "FALSE":
            sheet["A"+str(pos)].fill = PatternFill(start_color="FFFF00",end_color="FFFF00", fill_type="solid") # yellow
        elif is_in_last_month(odevr1):
            print("is last month:", odevr1)
            sheet["A"+str(pos)].fill = PatternFill(start_color="CCFFFF",end_color="CCFFFF", fill_type="solid") # light blue

        
        sheet["B"+str(pos)] = people_dict[odevr1]["leader"]

        sheet["C"+str(pos)] = people_dict[odevr2]["name"]
        if people_dict[odevr2]["has_experience"] == "FALSE":
            sheet["C"+str(pos)].fill = PatternFill(start_color="FFFF00",end_color="FFFF00", fill_type="solid") # yellow
        elif is_in_last_month(odevr2):
            print("is last month:", odevr2)
            sheet["C"+str(pos)].fill = PatternFill(start_color="CCFFFF",end_color="CCFFFF", fill_type="solid") # light blue

        sheet["D"+str(pos)] = people_dict[odevr2]["leader"]
        
        sheet["E"+str(pos)] = "=OR(D{pos}=D{npos},D{pos}=B{npos},B{pos}=B{npos},B{pos}=D{npos})".format(pos=pos,npos=(pos+1))

        ### New:
        sheet["F"+str(pos)] = people_dict[ndevr1]["name"]
        if people_dict[ndevr1]["has_experience"] == "FALSE":
            sheet["F"+str(pos)].fill = PatternFill(start_color="FFFF00",end_color="FFFF00", fill_type="solid") # yellow
        elif is_in_last_month(ndevr1):
            print("is last month:", ndevr1)
            sheet["F"+str(pos)].fill = PatternFill(start_color="CCFFFF",end_color="CCFFFF", fill_type="solid") # light blue

        sheet["G"+str(pos)] = people_dict[ndevr1]["leader"]
    
        sheet["H"+str(pos)] = people_dict[ndevr2]["name"]
        if people_dict[ndevr2]["has_experience"] == "FALSE":
            sheet["H"+str(pos)].fill = PatternFill(start_color="FFFF00",end_color="FFFF00", fill_type="solid") # yellow
        elif is_in_last_month(ndevr2):
            print("is last month:", ndevr2)
            sheet["H"+str(pos)].fill = PatternFill(start_color="CCFFFF",end_color="CCFFFF", fill_type="solid") # light blue

        sheet["I"+str(pos)] = people_dict[ndevr2]["leader"]

        sheet["J"+str(pos)]="=OR(I{pos}=I{npos},I{pos}=G{npos},G{pos}=G{npos},G{pos}=I{npos})".format(pos=pos,npos=(pos+1))
    
    offset += max(len(rota1), len(rota2) )+ 2
    sheet["A"+str(offset)] = "Code list"
    titles.append("A"+str(offset))
    offset +=1    

    for index, key in enumerate(people_dict):  
        pos = index + offset
        sheet["A"+str(pos)] = key
        sheet["B"+str(pos)] = people_dict[key]['name']
        sheet["C"+str(pos)] = people_dict[key]['leader']
    
    offset += len(people_dict) + 2

    sheet["A"+str(offset)] = "Last month"
    titles.append("A"+str(offset))
    offset +=1
    for index, key in enumerate(LAST_MONTH_L):
        pos = index + offset
        sheet["A"+str(pos)] = key
        sheet["B"+str(pos)] = people_dict[key]['name']

    
    for cell in titles:
        sheet[cell].alignment = Alignment(horizontal='center')
        sheet[cell].font = Font(bold=True)
    
    workbook.save("report.xlsx")
    print("Excel file created successfully.")

def write_solution_to_html(rota1, rota2, nrota1, nrota2, people_dict):
    model_to_print = []
    printI(self.best_solution)
    model_to_print.append("""
        var nodeDataArray = [
        {"isGroup":true,"key":"orig","text":"Original","xy":"0 0","width":300},
        {"key":"or1","group":"orig","text":"Rota 1", "isGroup":true},
        {"key":"or2","group":"orig","text":"Rota 2", "isGroup":true},
        {"isGroup":true,"key":"best","text":"Nueva","xy":"1500 0","width":300},
        {"key":"br1","group":"best","text":"Rota 1", "isGroup":true},
        {"key":"br2","group":"best","text":"Rota 2", "isGroup":true},
    """)
   
    nodeListToPrint = []
    linkToPrint = []

    for i, dev in enumerate(rota1):
        is_last_month =""
        if dev in LAST_MONTH_L:
            is_last_month = "*"
        nodeListToPrint.append('{"key":"o'+dev+'","group":"or1","text":"['+str(i)+'] '+dev[-1]+people_dict[dev]["name"]+' ('+people_dict[dev]['leader']+is_last_month+')"}')
        same_pos_r1 = i<len(nrota1) and dev == nrota1[i]
        same_pos_r2 = i<len(nrota2) and dev == nrota2[i]
        if not(same_pos_r1) and not(same_pos_r2):
            linkToPrint.append('{"from":"o'+dev+'","to":"b'+dev+'","category":"Mapping"}')

    for i, dev in enumerate(rota2):
        is_last_month =""
        if dev in LAST_MONTH_L:
            is_last_month = "*"
        nodeListToPrint.append('{"key":"o'+dev+'","group":"or2","text":"['+str(i)+'] '+dev[-1]+people_dict[dev]["name"]+' ('+people_dict[dev]['leader']+is_last_month+')"}')
        same_pos_r1 = i<len(nrota1) and dev == nrota1[i]
        same_pos_r2 = i<len(nrota2) and dev == nrota2[i]
        if not(same_pos_r1) and not(same_pos_r2):
            linkToPrint.append('{"from":"o'+dev+'","to":"b'+dev+'","category":"Mapping"}')

    for i, dev in enumerate(nrota1):
        is_last_month =""
        if dev in LAST_MONTH_L:
            is_last_month = "*"
        nodeListToPrint.append('{"key":"b'+dev+'","group":"br1","text":"['+str(i)+'] '+dev[-1]+people_dict[dev]["name"]+' ('+people_dict[dev]['leader']+is_last_month+')"}')

    for i, dev in enumerate(nrota2):
        is_last_month =""
        if dev in LAST_MONTH_L:
            is_last_month = "*"
        nodeListToPrint.append('{"key":"b'+dev+'","group":"br2","text":"['+str(i)+'] '+dev[-1]+people_dict[dev]["name"]+' ('+people_dict[dev]['leader']+is_last_month+')"}')
        
    model_to_print.append(",\n".join(nodeListToPrint))
    model_to_print.append('];\nvar linkDataArray = [')
                
    model_to_print.append(",\n".join(linkToPrint))

    model_to_print.append("]")
    replace_content("treeMapper_template.html","\n".join(model_to_print))    

def write_solution_to_file(rota1, rota2, nrota1, nrota2, people_dict):
    reportList = []
    reportList.append("New rotation")
    reportList.append("Original:" + (5*"\t") + "New:")
    reportList.append("Rota1\t\tRota2\t\t" + (1*"\t") + "Rota1\t\tRota2")
    reportList.append(("Name\tLeader\t"*2) + (1*"\t") + ("Name\tLeader\t"*2))
    for i in range(max(len(rota1), len(rota2))):
        
        odevr1, odevr2, ndevr1, ndevr2 = rota1[(i % len(rota1))], rota2[(i % len(rota2))], nrota1[(i % len(nrota1))], nrota2[(i % len(nrota2))]
        reportList.append("{}\t{}\t{}\t{}\t\t{}\t{}\t{}\t{}".format(
          
            ("* " if people_dict[odevr1]["has_experience"] == "FALSE" else "")+
            ("_ " if is_in_last_month(odevr1) else "")+
            people_dict[odevr1]["name"],
            people_dict[odevr1]["leader"],
          
            ("* " if people_dict[odevr2]["has_experience"] == "FALSE" else "")+
            ("_ " if is_in_last_month(odevr2) else "")+
            people_dict[odevr2]["name"],
            people_dict[odevr2]["leader"],
            
            ### New:
            ("* " if people_dict[ndevr1]["has_experience"] == "FALSE" else "")+
            ("_ " if is_in_last_month(ndevr1) else "")+
            people_dict[ndevr1]["name"],
            people_dict[ndevr1]["leader"],
       
            ("* " if people_dict[ndevr2]["has_experience"] == "FALSE" else "")+
            ("_ " if is_in_last_month(ndevr2) else "")+
            people_dict[ndevr2]["name"],
            people_dict[ndevr2]["leader"]
        ))

    report = "\n".join(reportList)

    with open("report.csv", "w") as file:
        file.write(report)
        

def replace_content(html_file, custom_text):
    """
    Replaces the placeholder ##CONTENT## with the specified custom text in an HTML file.
    """
    with open(html_file, "r") as file:
        html = file.read()
    html = html.replace("//##CONTENT##", custom_text)
    with open("treeMapper.html", "w") as file:
        file.write(html)

# Example usage:

if __name__ == '__main__':
    #csv_file = open('output_'+str(random.sample(range(88888), 1)[0])+'.csv', 'a')
    #writer = csv.writer(csv_file)
    print("open file with currentschedule")
    filename = 'output_'+str(random.sample(range(88888), 1)[0])+'.csv'
    print("filename: ", filename)
    import csv

    filename_input = "people.csv"  # Replace with the name of your CSV file
    leader_codes = {}  # Replace with the leader names and their codes

    # Load the CSV file into a list of dictionaries
    with open(filename_input, "r") as file:
        reader = csv.DictReader(file)
        people_list = list(reader)

    # Convert the list of people into a dictionary with codes as keys

    # Load the CSV file into a list of dictionaries
    with open("last_month.csv", "r") as file:
        reader = csv.DictReader(file)
        last_month = list(reader)
    
    devs = []
    dev_by_name = {}
    for i, person in enumerate(people_list):
        if not (leader_codes.get(person["leader"]) != None):
            leader_codes[person["leader"]] = chr(len(leader_codes) + 65)
        has_experience = person["has_experience"] == "TRUE" or person["name"] in last_month
        code =  "{:02d}{}{}{}".format(
                i,
                leader_codes[person["leader"]],
                str(i+1),
                ("_" if has_experience else "x")
        )
        people_dict[code] = person
        dev_by_name[person["name"]] = code
        devs.append(code)

    print(people_dict)

    original_devs_arrange = "-".join(devs)
    half_point = int(len(devs)/2)
   
    for lm in last_month:
        LAST_MONTH_L.append(dev_by_name[lm["name"]])

    print(LAST_MONTH_L)


    ga = EightQueensGA()
    #ga.set_csv_writer(writer)
    try:
        #sol = ['A15_', 'B12_', 'E28x', 'A42_', 'A14_', 'B08_', 'I47_', 'B17_', 'C06_', 'F37_', 'D10x', 'E11_', 'B21_', 'F30_', 'A01_', 'C20_', 'D07_', 'I35_', 'I46_', 'A04_', 'G26_', 'F22_', 'B02_','G24_', 'C13_', 'H31_', 'D27x', 'E09_', 'C03_', 'F40_', 'I36_', 'A19_', 'E25x', 'D44_', 'H32_', 'E43_', 'C05_', 'C16_', 'F38_', 'A18_', 'H29_', 'H45_', 'C33_', 'F39_', 'I41_', 'E23x', 'I34_']
        #print(fitness(sol))
        #printI(sol)
        print("Initial Max childs {} then {} are random".format(MAX_CHILDS, POPULATION_SIZE-MAX_CHILDS))
        ga.evolve()
        ga.print_solution()  
    except KeyboardInterrupt:
        ga.print_solution()
        #csv_file.close()
        sys.exit(0)


